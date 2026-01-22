"""
PO Breakdown Export and Reporting Service

This service provides comprehensive export and reporting functionality for SAP PO Breakdown data,
supporting multiple formats (CSV, Excel, JSON) with hierarchical relationship preservation.
"""

import csv
import json
from datetime import datetime
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional
from uuid import UUID

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from supabase import Client


class POBreakdownExportService:
    """Service for exporting PO breakdown data in multiple formats"""
    
    def __init__(self, supabase: Client):
        self.supabase = supabase
    
    async def export_to_csv(
        self,
        project_id: UUID,
        filters: Optional[Dict[str, Any]] = None,
        include_hierarchy: bool = True,
        custom_fields: Optional[List[str]] = None
    ) -> str:
        """
        Export PO breakdown data to CSV format
        
        Args:
            project_id: Project UUID
            filters: Optional filters to apply
            include_hierarchy: Whether to include hierarchy indicators
            custom_fields: List of custom field keys to include
            
        Returns:
            CSV string
        """
        # Fetch data
        breakdowns = await self._fetch_breakdowns(project_id, filters)
        
        if not breakdowns:
            return ""
        
        # Prepare CSV
        output = StringIO()
        
        # Define base columns
        base_columns = [
            'id', 'name', 'code', 'sap_po_number', 'sap_line_item',
            'hierarchy_level', 'parent_breakdown_id', 'cost_center', 'gl_account',
            'planned_amount', 'committed_amount', 'actual_amount', 'remaining_amount',
            'currency', 'breakdown_type', 'category', 'subcategory',
            'tags', 'notes', 'created_at', 'updated_at'
        ]
        
        # Add custom field columns if specified
        if custom_fields:
            base_columns.extend([f'custom_{field}' for field in custom_fields])
        
        writer = csv.DictWriter(output, fieldnames=base_columns, extrasaction='ignore')
        writer.writeheader()
        
        # Write data rows
        for breakdown in breakdowns:
            row = self._prepare_csv_row(breakdown, include_hierarchy, custom_fields)
            writer.writerow(row)
        
        return output.getvalue()
    
    async def export_to_excel(
        self,
        project_id: UUID,
        filters: Optional[Dict[str, Any]] = None,
        include_hierarchy: bool = True,
        include_summary: bool = True,
        custom_fields: Optional[List[str]] = None
    ) -> BytesIO:
        """
        Export PO breakdown data to Excel format with formatting
        
        Args:
            project_id: Project UUID
            filters: Optional filters to apply
            include_hierarchy: Whether to include hierarchy indicators
            include_summary: Whether to include summary sheet
            custom_fields: List of custom field keys to include
            
        Returns:
            BytesIO containing Excel file
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        # Fetch data
        breakdowns = await self._fetch_breakdowns(project_id, filters)
        
        # Create workbook
        wb = Workbook()
        
        # Create data sheet
        ws_data = wb.active
        ws_data.title = "PO Breakdowns"
        
        # Define columns
        base_columns = [
            ('ID', 'id'),
            ('Name', 'name'),
            ('Code', 'code'),
            ('SAP PO Number', 'sap_po_number'),
            ('SAP Line Item', 'sap_line_item'),
            ('Hierarchy Level', 'hierarchy_level'),
            ('Parent ID', 'parent_breakdown_id'),
            ('Cost Center', 'cost_center'),
            ('GL Account', 'gl_account'),
            ('Planned Amount', 'planned_amount'),
            ('Committed Amount', 'committed_amount'),
            ('Actual Amount', 'actual_amount'),
            ('Remaining Amount', 'remaining_amount'),
            ('Currency', 'currency'),
            ('Type', 'breakdown_type'),
            ('Category', 'category'),
            ('Subcategory', 'subcategory'),
            ('Tags', 'tags'),
            ('Notes', 'notes'),
            ('Created At', 'created_at'),
            ('Updated At', 'updated_at')
        ]
        
        # Add custom field columns
        if custom_fields:
            base_columns.extend([(f'Custom: {field}', f'custom_{field}') for field in custom_fields])
        
        # Write headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, (header, _) in enumerate(base_columns, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data rows
        for row_idx, breakdown in enumerate(breakdowns, 2):
            for col_idx, (_, field) in enumerate(base_columns, 1):
                value = self._get_field_value(breakdown, field, custom_fields)
                
                # Apply hierarchy indentation for name column
                if field == 'name' and include_hierarchy:
                    level = breakdown.get('hierarchy_level', 0)
                    value = '  ' * level + str(value)
                
                cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
                
                # Format numeric columns
                if field in ['planned_amount', 'committed_amount', 'actual_amount', 'remaining_amount']:
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for col_idx in range(1, len(base_columns) + 1):
            ws_data.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Create summary sheet if requested
        if include_summary:
            ws_summary = wb.create_sheet("Summary")
            await self._create_summary_sheet(ws_summary, breakdowns)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    async def export_to_json(
        self,
        project_id: UUID,
        filters: Optional[Dict[str, Any]] = None,
        include_hierarchy: bool = True,
        hierarchical_structure: bool = False
    ) -> str:
        """
        Export PO breakdown data to JSON format
        
        Args:
            project_id: Project UUID
            filters: Optional filters to apply
            include_hierarchy: Whether to include hierarchy metadata
            hierarchical_structure: Whether to nest children under parents
            
        Returns:
            JSON string
        """
        # Fetch data
        breakdowns = await self._fetch_breakdowns(project_id, filters)
        
        if hierarchical_structure:
            # Build hierarchical structure
            data = self._build_hierarchical_json(breakdowns)
        else:
            # Flat structure
            data = [self._prepare_json_item(b) for b in breakdowns]
        
        # Add metadata
        export_data = {
            'export_timestamp': datetime.now().isoformat(),
            'project_id': str(project_id),
            'total_items': len(breakdowns),
            'filters_applied': filters or {},
            'data': data
        }
        
        return json.dumps(export_data, indent=2, default=str)
    
    async def generate_financial_summary(
        self,
        project_id: UUID,
        group_by: Optional[str] = None,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Generate financial summary report with aggregations
        
        Args:
            project_id: Project UUID
            group_by: Field to group by (category, hierarchy_level, cost_center)
            filters: Optional filters to apply
            
        Returns:
            Dictionary containing summary data
        """
        # Fetch data
        breakdowns = await self._fetch_breakdowns(project_id, filters)
        
        if not breakdowns:
            return {
                'project_id': str(project_id),
                'total_items': 0,
                'totals': self._empty_totals(),
                'groups': {}
            }
        
        # Calculate overall totals
        totals = self._calculate_totals(breakdowns)
        
        # Calculate grouped totals if requested
        groups = {}
        if group_by:
            groups = self._calculate_grouped_totals(breakdowns, group_by)
        
        return {
            'project_id': str(project_id),
            'generated_at': datetime.now().isoformat(),
            'total_items': len(breakdowns),
            'totals': totals,
            'groups': groups,
            'filters_applied': filters or {}
        }
    
    async def _fetch_breakdowns(
        self,
        project_id: UUID,
        filters: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """Fetch PO breakdowns with optional filters"""
        
        query = self.supabase.table('po_breakdowns').select('*').eq('project_id', str(project_id)).eq('is_active', True)
        
        # Apply filters
        if filters:
            if 'breakdown_type' in filters:
                query = query.eq('breakdown_type', filters['breakdown_type'])
            
            if 'cost_center' in filters:
                query = query.eq('cost_center', filters['cost_center'])
            
            if 'category' in filters:
                query = query.eq('category', filters['category'])
            
            if 'min_planned_amount' in filters:
                query = query.gte('planned_amount', str(filters['min_planned_amount']))
            
            if 'max_planned_amount' in filters:
                query = query.lte('planned_amount', str(filters['max_planned_amount']))
            
            if 'search' in filters:
                # Note: This is a simplified search, real implementation would use full-text search
                query = query.ilike('name', f"%{filters['search']}%")
        
        # Order by hierarchy
        query = query.order('hierarchy_level').order('name')
        
        result = query.execute()
        breakdowns = result.data or []
        
        # Calculate remaining amounts
        for breakdown in breakdowns:
            planned = Decimal(str(breakdown.get('planned_amount', 0)))
            actual = Decimal(str(breakdown.get('actual_amount', 0)))
            breakdown['remaining_amount'] = str(planned - actual)
        
        return breakdowns
    
    def _prepare_csv_row(
        self,
        breakdown: Dict[str, Any],
        include_hierarchy: bool,
        custom_fields: Optional[List[str]]
    ) -> Dict[str, Any]:
        """Prepare a breakdown item for CSV export"""
        
        row = {
            'id': breakdown.get('id', ''),
            'name': breakdown.get('name', ''),
            'code': breakdown.get('code', ''),
            'sap_po_number': breakdown.get('sap_po_number', ''),
            'sap_line_item': breakdown.get('sap_line_item', ''),
            'hierarchy_level': breakdown.get('hierarchy_level', 0),
            'parent_breakdown_id': breakdown.get('parent_breakdown_id', ''),
            'cost_center': breakdown.get('cost_center', ''),
            'gl_account': breakdown.get('gl_account', ''),
            'planned_amount': breakdown.get('planned_amount', 0),
            'committed_amount': breakdown.get('committed_amount', 0),
            'actual_amount': breakdown.get('actual_amount', 0),
            'remaining_amount': breakdown.get('remaining_amount', 0),
            'currency': breakdown.get('currency', 'USD'),
            'breakdown_type': breakdown.get('breakdown_type', ''),
            'category': breakdown.get('category', ''),
            'subcategory': breakdown.get('subcategory', ''),
            'tags': ','.join(breakdown.get('tags', [])),
            'notes': breakdown.get('notes', ''),
            'created_at': breakdown.get('created_at', ''),
            'updated_at': breakdown.get('updated_at', '')
        }
        
        # Add hierarchy indentation to name if requested
        if include_hierarchy:
            level = breakdown.get('hierarchy_level', 0)
            row['name'] = '  ' * level + row['name']
        
        # Add custom fields
        if custom_fields:
            custom_data = breakdown.get('custom_fields', {})
            for field in custom_fields:
                row[f'custom_{field}'] = custom_data.get(field, '')
        
        return row
    
    def _get_field_value(
        self,
        breakdown: Dict[str, Any],
        field: str,
        custom_fields: Optional[List[str]]
    ) -> Any:
        """Get field value from breakdown, handling custom fields"""
        
        if field.startswith('custom_') and custom_fields:
            custom_key = field.replace('custom_', '')
            custom_data = breakdown.get('custom_fields', {})
            return custom_data.get(custom_key, '')
        
        if field == 'tags':
            return ','.join(breakdown.get('tags', []))
        
        return breakdown.get(field, '')
    
    def _prepare_json_item(self, breakdown: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare a breakdown item for JSON export"""
        
        return {
            'id': breakdown.get('id'),
            'name': breakdown.get('name'),
            'code': breakdown.get('code'),
            'sap_po_number': breakdown.get('sap_po_number'),
            'sap_line_item': breakdown.get('sap_line_item'),
            'hierarchy_level': breakdown.get('hierarchy_level', 0),
            'parent_breakdown_id': breakdown.get('parent_breakdown_id'),
            'cost_center': breakdown.get('cost_center'),
            'gl_account': breakdown.get('gl_account'),
            'financial_data': {
                'planned_amount': str(breakdown.get('planned_amount', 0)),
                'committed_amount': str(breakdown.get('committed_amount', 0)),
                'actual_amount': str(breakdown.get('actual_amount', 0)),
                'remaining_amount': str(breakdown.get('remaining_amount', 0)),
                'currency': breakdown.get('currency', 'USD')
            },
            'breakdown_type': breakdown.get('breakdown_type'),
            'category': breakdown.get('category'),
            'subcategory': breakdown.get('subcategory'),
            'custom_fields': breakdown.get('custom_fields', {}),
            'tags': breakdown.get('tags', []),
            'notes': breakdown.get('notes'),
            'metadata': {
                'import_batch_id': breakdown.get('import_batch_id'),
                'import_source': breakdown.get('import_source'),
                'version': breakdown.get('version', 1),
                'created_at': breakdown.get('created_at'),
                'updated_at': breakdown.get('updated_at')
            }
        }
    
    def _build_hierarchical_json(self, breakdowns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Build hierarchical JSON structure with nested children"""
        
        # Create lookup dictionary
        breakdown_dict = {b['id']: self._prepare_json_item(b) for b in breakdowns}
        
        # Add children arrays
        for item in breakdown_dict.values():
            item['children'] = []
        
        # Build hierarchy
        root_items = []
        for breakdown in breakdowns:
            item_id = breakdown['id']
            parent_id = breakdown.get('parent_breakdown_id')
            
            if parent_id and parent_id in breakdown_dict:
                # Add to parent's children
                breakdown_dict[parent_id]['children'].append(breakdown_dict[item_id])
            else:
                # Root level item
                root_items.append(breakdown_dict[item_id])
        
        return root_items
    
    def _calculate_totals(self, breakdowns: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate financial totals"""
        
        total_planned = Decimal('0')
        total_committed = Decimal('0')
        total_actual = Decimal('0')
        total_remaining = Decimal('0')
        
        for breakdown in breakdowns:
            total_planned += Decimal(str(breakdown.get('planned_amount', 0)))
            total_committed += Decimal(str(breakdown.get('committed_amount', 0)))
            total_actual += Decimal(str(breakdown.get('actual_amount', 0)))
            total_remaining += Decimal(str(breakdown.get('remaining_amount', 0)))
        
        # Calculate variances
        variance_amount = total_planned - total_actual
        variance_percentage = (variance_amount / total_planned * 100) if total_planned > 0 else Decimal('0')
        
        return {
            'planned_amount': str(total_planned),
            'committed_amount': str(total_committed),
            'actual_amount': str(total_actual),
            'remaining_amount': str(total_remaining),
            'variance_amount': str(variance_amount),
            'variance_percentage': str(round(variance_percentage, 2)),
            'currency': breakdowns[0].get('currency', 'USD') if breakdowns else 'USD'
        }
    
    def _calculate_grouped_totals(
        self,
        breakdowns: List[Dict[str, Any]],
        group_by: str
    ) -> Dict[str, Dict[str, Any]]:
        """Calculate totals grouped by specified field"""
        
        groups = {}
        
        for breakdown in breakdowns:
            group_key = breakdown.get(group_by, 'Unknown')
            if group_key is None:
                group_key = 'Unknown'
            
            group_key = str(group_key)
            
            if group_key not in groups:
                groups[group_key] = []
            
            groups[group_key].append(breakdown)
        
        # Calculate totals for each group
        result = {}
        for group_key, group_breakdowns in groups.items():
            result[group_key] = {
                'count': len(group_breakdowns),
                'totals': self._calculate_totals(group_breakdowns)
            }
        
        return result
    
    def _empty_totals(self) -> Dict[str, Any]:
        """Return empty totals structure"""
        
        return {
            'planned_amount': '0',
            'committed_amount': '0',
            'actual_amount': '0',
            'remaining_amount': '0',
            'variance_amount': '0',
            'variance_percentage': '0',
            'currency': 'USD'
        }
    
    async def _create_summary_sheet(self, worksheet, breakdowns: List[Dict[str, Any]]):
        """Create summary sheet in Excel workbook"""
        
        # Title
        worksheet['A1'] = 'PO Breakdown Summary Report'
        worksheet['A1'].font = Font(bold=True, size=14)
        
        # Generation timestamp
        worksheet['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # Overall totals
        worksheet['A4'] = 'Overall Totals'
        worksheet['A4'].font = Font(bold=True)
        
        totals = self._calculate_totals(breakdowns)
        
        worksheet['A5'] = 'Total Items:'
        worksheet['B5'] = len(breakdowns)
        
        worksheet['A6'] = 'Planned Amount:'
        worksheet['B6'] = float(totals['planned_amount'])
        worksheet['B6'].number_format = '#,##0.00'
        
        worksheet['A7'] = 'Committed Amount:'
        worksheet['B7'] = float(totals['committed_amount'])
        worksheet['B7'].number_format = '#,##0.00'
        
        worksheet['A8'] = 'Actual Amount:'
        worksheet['B8'] = float(totals['actual_amount'])
        worksheet['B8'].number_format = '#,##0.00'
        
        worksheet['A9'] = 'Remaining Amount:'
        worksheet['B9'] = float(totals['remaining_amount'])
        worksheet['B9'].number_format = '#,##0.00'
        
        worksheet['A10'] = 'Variance:'
        worksheet['B10'] = float(totals['variance_amount'])
        worksheet['B10'].number_format = '#,##0.00'
        
        worksheet['A11'] = 'Variance %:'
        worksheet['B11'] = float(totals['variance_percentage'])
        worksheet['B11'].number_format = '0.00%'
        
        # By category
        worksheet['A13'] = 'Breakdown by Category'
        worksheet['A13'].font = Font(bold=True)
        
        category_groups = self._calculate_grouped_totals(breakdowns, 'category')
        
        row = 14
        for category, data in category_groups.items():
            worksheet[f'A{row}'] = category
            worksheet[f'B{row}'] = data['count']
            worksheet[f'C{row}'] = float(data['totals']['planned_amount'])
            worksheet[f'C{row}'].number_format = '#,##0.00'
            row += 1
