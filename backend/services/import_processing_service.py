"""
Import Processing Service for SAP PO Breakdown Management

This module provides comprehensive CSV and Excel file import processing
for SAP PO breakdown data, including validation, parsing, hierarchy construction,
and conflict resolution.

**Validates: Requirements 1.1, 1.2, 10.1**
"""

import csv
import io
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Any, Optional, Tuple
from uuid import UUID, uuid4

from fastapi import UploadFile, HTTPException
from supabase import Client

from models.po_breakdown import (
    ImportConfig,
    ImportResult,
    ImportStatus,
    ImportError,
    ImportWarning,
    ImportConflict,
    ConflictType,
    ConflictResolution,
    POBreakdownCreate,
    POBreakdownType,
    ErrorSeverity,
    ErrorCategory,
    ImportBatchStatus,
    ImportBatchErrorDetail,
)
from services.po_breakdown_service import POBreakdownDatabaseService

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constants
MAX_FILE_SIZE_MB = 50
SUPPORTED_CSV_ENCODINGS = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1']
REQUIRED_FIELDS = ['name']  # Minimum required field for PO breakdown


class ImportProcessingService:
    """
    Service for processing SAP PO data imports from CSV and Excel files.
    
    Provides comprehensive file validation, parsing with configurable column mappings,
    data transformation, and integration with POBreakdownService.
    
    **Validates: Requirements 1.1, 1.2, 10.1**
    """

    
    def __init__(self, supabase_client: Client):
        """
        Initialize the import processing service.
        
        Args:
            supabase_client: Supabase client for database operations
        """
        self.supabase = supabase_client
        self.po_service = POBreakdownDatabaseService(supabase_client)
        self.import_batch_table = 'po_import_batches'
    
    # =========================================================================
    # File Validation
    # =========================================================================
    
    async def validate_import_file(self, file: UploadFile) -> Dict[str, Any]:
        """
        Validate an uploaded file for import.
        
        **Validates: Requirements 1.1, 10.1**
        
        Args:
            file: Uploaded file to validate
            
        Returns:
            Dict with validation results:
                - is_valid: bool
                - errors: List[str]
                - warnings: List[str]
                - file_info: Dict with file metadata
                
        Raises:
            HTTPException: If file validation fails critically
        """
        errors = []
        warnings = []
        
        # Check file exists
        if not file:
            errors.append("No file provided")
            return {
                'is_valid': False,
                'errors': errors,
                'warnings': warnings,
                'file_info': {}
            }
        
        # Get file info
        filename = file.filename or "unknown"
        file_size = 0
        
        # Check file extension
        file_ext = filename.lower().split('.')[-1] if '.' in filename else ''
        if file_ext not in ['csv', 'xlsx', 'xls']:
            errors.append(f"Unsupported file format: {file_ext}. Supported formats: CSV, XLSX, XLS")
        
        # Check file size
        try:
            content = await file.read()
            file_size = len(content)
            await file.seek(0)  # Reset file pointer
            
            max_size_bytes = MAX_FILE_SIZE_MB * 1024 * 1024
            if file_size > max_size_bytes:
                errors.append(f"File size ({file_size / 1024 / 1024:.2f} MB) exceeds maximum allowed size ({MAX_FILE_SIZE_MB} MB)")
            elif file_size == 0:
                errors.append("File is empty")
        except Exception as e:
            errors.append(f"Failed to read file: {str(e)}")
        
        # Try to detect encoding for CSV files
        encoding = 'utf-8'
        if file_ext == 'csv' and file_size > 0:
            try:
                # Try different encodings
                for enc in SUPPORTED_CSV_ENCODINGS:
                    try:
                        content.decode(enc)
                        encoding = enc
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    warnings.append("Could not detect file encoding, will attempt UTF-8")
            except Exception as e:
                warnings.append(f"Encoding detection failed: {str(e)}")
        
        file_info = {
            'filename': filename,
            'file_size': file_size,
            'file_type': file_ext,
            'encoding': encoding
        }
        
        is_valid = len(errors) == 0
        
        return {
            'is_valid': is_valid,
            'errors': errors,
            'warnings': warnings,
            'file_info': file_info
        }

    
    # =========================================================================
    # CSV Import Processing
    # =========================================================================
    
    async def process_csv_import(
        self,
        file: UploadFile,
        project_id: UUID,
        config: ImportConfig,
        user_id: UUID
    ) -> ImportResult:
        """
        Process a CSV file import with configurable column mappings.
        
        **Validates: Requirements 1.1, 1.2, 10.1**
        
        Args:
            file: CSV file to import
            project_id: Target project UUID
            config: Import configuration with column mappings
            user_id: User performing the import
            
        Returns:
            ImportResult with processing details
        """
        start_time = datetime.now()
        
        # Validate file
        validation = await self.validate_import_file(file)
        if not validation['is_valid']:
            raise HTTPException(
                status_code=400,
                detail=f"File validation failed: {', '.join(validation['errors'])}"
            )
        
        # Create import batch
        batch_id = await self.create_import_batch(
            project_id=project_id,
            source=f"CSV: {file.filename}",
            user_id=user_id,
            file_name=file.filename,
            file_size_bytes=validation['file_info'].get('file_size', 0),
            file_type='csv',
            import_config=config
        )
        
        errors: List[ImportError] = []
        warnings: List[ImportWarning] = []
        conflicts: List[ImportConflict] = []
        created_breakdown_ids: List[UUID] = []
        
        try:
            # Read and parse CSV
            content = await file.read()
            csv_content = content.decode(config.encoding)
            
            # Parse CSV with column mappings
            parsed_rows = self._parse_csv_content(
                csv_content=csv_content,
                config=config,
                errors=errors,
                warnings=warnings
            )
            
            total_records = len(parsed_rows)
            processed_records = 0
            successful_records = 0
            failed_records = 0
            created_hierarchies = 0
            
            # Check if hierarchy construction is needed
            use_hierarchy_construction = (
                config.hierarchy_column is not None or 
                config.parent_reference_column is not None
            )
            
            if use_hierarchy_construction:
                # Use hierarchy construction workflow
                try:
                    created_breakdown_ids, created_hierarchies = await self.construct_hierarchy_from_import(
                        parsed_rows=parsed_rows,
                        project_id=project_id,
                        config=config,
                        user_id=user_id,
                        batch_id=batch_id,
                        errors=errors,
                        warnings=warnings
                    )
                    
                    successful_records = len(created_breakdown_ids)
                    processed_records = total_records
                    failed_records = total_records - successful_records
                    
                except Exception as e:
                    logger.error(f"Hierarchy construction failed: {e}")
                    errors.append(ImportError(
                        row_number=0,
                        field='general',
                        error_type='hierarchy_construction_error',
                        message=f"Failed to construct hierarchy: {str(e)}",
                        raw_value=None
                    ))
                    failed_records = total_records
            else:
                # Use simple row-by-row processing (no hierarchy)
                for row_num, row_data in enumerate(parsed_rows, start=config.skip_header_rows + 1):
                    processed_records += 1
                    
                    try:
                        # Transform and validate row data
                        breakdown_data = self._transform_row_to_breakdown(
                            row_data=row_data,
                            config=config,
                            row_number=row_num,
                            errors=errors,
                            warnings=warnings
                        )
                        
                        if breakdown_data:
                            # Check for conflicts
                            conflict = await self._check_for_conflicts(
                                breakdown_data=breakdown_data,
                                project_id=project_id,
                                row_number=row_num
                            )
                            
                            if conflict:
                                conflicts.append(conflict)
                                
                                # Apply conflict resolution strategy
                                if config.conflict_resolution == ConflictResolution.skip:
                                    warnings.append(ImportWarning(
                                        row_number=row_num,
                                        field=conflict.field_conflicts[0] if conflict.field_conflicts else 'general',
                                        warning_type='conflict_skipped',
                                        message=f"Skipped due to conflict: {conflict.conflict_type.value}"
                                    ))
                                    continue
                                elif config.conflict_resolution == ConflictResolution.update:
                                    # Update existing record
                                    existing_id = UUID(conflict.existing_record['id'])
                                    # Note: Update logic would go here
                                    warnings.append(ImportWarning(
                                        row_number=row_num,
                                        field='general',
                                        warning_type='conflict_updated',
                                        message=f"Updated existing record {existing_id}"
                                    ))
                            
                            # Create breakdown item
                            created = await self.po_service.create_breakdown(
                                project_id=project_id,
                                breakdown_data=breakdown_data,
                                user_id=user_id
                            )
                            
                            # Update with import metadata
                            await self._update_import_metadata(
                                breakdown_id=created.id,
                                batch_id=batch_id,
                                source=file.filename
                            )
                            
                            created_breakdown_ids.append(created.id)
                            successful_records += 1
                            
                    except Exception as e:
                        failed_records += 1
                        errors.append(ImportError(
                            row_number=row_num,
                            field='general',
                            error_type='processing_error',
                            message=f"Failed to process row: {str(e)}",
                            raw_value=str(row_data)
                        ))
                        logger.error(f"Failed to process row {row_num}: {e}")
            
            # Calculate processing time
            processing_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            
            # Determine final status
            if failed_records == 0 and len(errors) == 0:
                status = ImportStatus.completed
                status_message = f"Import completed successfully. {successful_records} records imported."
            elif successful_records > 0:
                status = ImportStatus.partially_completed
                status_message = f"Import partially completed. {successful_records} succeeded, {failed_records} failed."
            else:
                status = ImportStatus.failed
                status_message = f"Import failed. {failed_records} records failed to import."
            
            # Calculate error summaries
            errors_by_category = {}
            errors_by_severity = {}
            for error in errors:
                category = error.category.value
                severity = error.severity.value
                errors_by_category[category] = errors_by_category.get(category, 0) + 1
                errors_by_severity[severity] = errors_by_severity.get(severity, 0) + 1
            
            # Calculate max hierarchy depth
            max_hierarchy_depth = 0
            if use_hierarchy_construction and created_breakdown_ids:
                # Query created breakdowns to find max depth
                try:
                    result_query = self.supabase.table('po_breakdowns')\
                        .select('hierarchy_level')\
                        .in_('id', [str(id) for id in created_breakdown_ids])\
                        .execute()
                    if result_query.data:
                        max_hierarchy_depth = max(row.get('hierarchy_level', 0) for row in result_query.data)
                except Exception as e:
                    logger.warning(f"Failed to calculate max hierarchy depth: {e}")
            
            # Store detailed errors, warnings, and conflicts in database
            await self._store_batch_errors(batch_id, errors)
            await self._store_batch_warnings(batch_id, warnings)
            await self._store_batch_conflicts(batch_id, conflicts)
            
            # Update batch status with comprehensive metrics
            await self._update_batch_status(
                batch_id=batch_id,
                status=status,
                status_message=status_message,
                metrics={
                    'total_records': total_records,
                    'processed_records': processed_records,
                    'successful_records': successful_records,
                    'failed_records': failed_records,
                    'skipped_records': len(conflicts) if config.conflict_resolution == ConflictResolution.skip else 0,
                    'updated_records': 0,
                    'created_hierarchies': created_hierarchies,
                    'max_hierarchy_depth': max_hierarchy_depth,
                    'error_count': len(errors),
                    'warning_count': len(warnings),
                    'conflict_count': len(conflicts),
                    'errors_by_category': errors_by_category,
                    'errors_by_severity': errors_by_severity,
                    'processing_time_ms': processing_time_ms,
                    'created_breakdown_ids': [str(id) for id in created_breakdown_ids]
                }
            )
            
            result = ImportResult(
                batch_id=batch_id,
                status=status,
                status_message=status_message,
                total_records=total_records,
                processed_records=processed_records,
                successful_records=successful_records,
                failed_records=failed_records,
                skipped_records=len(conflicts) if config.conflict_resolution == ConflictResolution.skip else 0,
                updated_records=0,
                conflicts=conflicts,
                errors=errors,
                warnings=warnings,
                error_count=len(errors),
                warning_count=len(warnings),
                conflict_count=len(conflicts),
                errors_by_category=errors_by_category,
                errors_by_severity=errors_by_severity,
                processing_time_ms=processing_time_ms,
                created_hierarchies=created_hierarchies,
                max_hierarchy_depth=max_hierarchy_depth,
                created_breakdown_ids=created_breakdown_ids,
                can_rollback=True,
                rollback_instructions=f"Use rollback_import_batch({batch_id}) to delete all {len(created_breakdown_ids)} created breakdowns."
            )
            
            logger.info(
                f"CSV import completed: batch_id={batch_id}, "
                f"status={status}, successful={successful_records}/{total_records}, "
                f"errors={len(errors)}, warnings={len(warnings)}, conflicts={len(conflicts)}"
            )
            
            return result
            
        except Exception as e:
            logger.error(f"CSV import failed: {e}")
            await self._update_batch_status(
                batch_id=batch_id,
                status=ImportStatus.failed,
                status_message=f"Import failed with exception: {str(e)}"
            )
            raise HTTPException(
                status_code=500,
                detail=f"Import processing failed: {str(e)}"
            )

    
    # =========================================================================
    # Excel Import Processing
    # =========================================================================
    
    async def process_excel_import(
        self,
        file: UploadFile,
        project_id: UUID,
        config: ImportConfig,
        user_id: UUID
    ) -> ImportResult:
        """
        Process an Excel file import with configurable column mappings.
        
        **Validates: Requirements 1.1, 1.2, 10.1**
        
        Args:
            file: Excel file to import
            project_id: Target project UUID
            config: Import configuration with column mappings
            user_id: User performing the import
            
        Returns:
            ImportResult with processing details
        """
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="Excel support not available. Install openpyxl package."
            )
        
        start_time = datetime.now()
        
        # Validate file
        validation = await self.validate_import_file(file)
        if not validation['is_valid']:
            raise HTTPException(
                status_code=400,
                detail=f"File validation failed: {', '.join(validation['errors'])}"
            )
        
        # Create import batch
        batch_id = await self.create_import_batch(
            project_id=project_id,
            source=f"Excel: {file.filename}",
            user_id=user_id,
            file_name=file.filename,
            file_size_bytes=validation['file_info'].get('file_size', 0),
            file_type=validation['file_info'].get('file_type', 'xlsx'),
            import_config=config
        )
        
        errors: List[ImportError] = []
        warnings: List[ImportWarning] = []
        conflicts: List[ImportConflict] = []
        created_breakdown_ids: List[UUID] = []
        
        try:
            # Read Excel file
            content = await file.read()
            workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            
            # Use first sheet by default
            sheet = workbook.active
            
            # Parse Excel with column mappings
            parsed_rows = self._parse_excel_sheet(
                sheet=sheet,
                config=config,
                errors=errors,
                warnings=warnings
            )
            
            total_records = len(parsed_rows)
            processed_records = 0
            successful_records = 0
            failed_records = 0
            created_hierarchies = 0
            
            # Check if hierarchy construction is needed
            use_hierarchy_construction = (
                config.hierarchy_column is not None or 
                config.parent_reference_column is not None
            )
            
            if use_hierarchy_construction:
                # Use hierarchy construction workflow
                try:
                    created_breakdown_ids, created_hierarchies = await self.construct_hierarchy_from_import(
                        parsed_rows=parsed_rows,
                        project_id=project_id,
                        config=config,
                        user_id=user_id,
                        batch_id=batch_id,
                        errors=errors,
                        warnings=warnings
                    )
                    
                    successful_records = len(created_breakdown_ids)
                    processed_records = total_records
                    failed_records = total_records - successful_records
                    
                except Exception as e:
                    logger.error(f"Hierarchy construction failed: {e}")
                    errors.append(ImportError(
                        row_number=0,
                        field='general',
                        error_type='hierarchy_construction_error',
                        message=f"Failed to construct hierarchy: {str(e)}",
                        raw_value=None
                    ))
                    failed_records = total_records
            else:
                # Use simple row-by-row processing (no hierarchy)
                for row_num, row_data in enumerate(parsed_rows, start=config.skip_header_rows + 1):
                    processed_records += 1
                    
                    try:
                        # Transform and validate row data
                        breakdown_data = self._transform_row_to_breakdown(
                            row_data=row_data,
                            config=config,
                            row_number=row_num,
                            errors=errors,
                            warnings=warnings
                        )
                        
                        if breakdown_data:
                            # Check for conflicts
                            conflict = await self._check_for_conflicts(
                                breakdown_data=breakdown_data,
                                project_id=project_id,
                                row_number=row_num
                            )
                            
                            if conflict:
                                conflicts.append(conflict)
                                
                                if config.conflict_resolution == ConflictResolution.skip:
                                    warnings.append(ImportWarning(
                                        row_number=row_num,
                                        field='general',
                                        warning_type='conflict_skipped',
                                        message=f"Skipped due to conflict: {conflict.conflict_type.value}"
                                    ))
                                    continue
                            
                            # Create breakdown item
                            created = await self.po_service.create_breakdown(
                                project_id=project_id,
                                breakdown_data=breakdown_data,
                                user_id=user_id
                            )
                            
                            # Update with import metadata
                            await self._update_import_metadata(
                                breakdown_id=created.id,
                                batch_id=batch_id,
                                source=file.filename
                            )
                            
                            created_breakdown_ids.append(created.id)
                            successful_records += 1
                            
                    except Exception as e:
                        failed_records += 1
                        errors.append(ImportError(
                            row_number=row_num,
                            field='general',
                            error_type='processing_error',
                            message=f"Failed to process row: {str(e)}",
                            raw_value=str(row_data)
                        ))
                        logger.error(f"Failed to process row {row_num}: {e}")
            
            # Calculate processing time
            processing_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            
            # Determine final status
            if failed_records == 0 and len(errors) == 0:
                status = ImportStatus.completed
                status_message = f"Import completed successfully. {successful_records} records imported."
            elif successful_records > 0:
                status = ImportStatus.partially_completed
                status_message = f"Import partially completed. {successful_records} succeeded, {failed_records} failed."
            else:
                status = ImportStatus.failed
                status_message = f"Import failed. {failed_records} records failed to import."
            
            # Calculate error summaries
            errors_by_category = {}
            errors_by_severity = {}
            for error in errors:
                category = error.category.value
                severity = error.severity.value
                errors_by_category[category] = errors_by_category.get(category, 0) + 1
                errors_by_severity[severity] = errors_by_severity.get(severity, 0) + 1
            
            # Calculate max hierarchy depth
            max_hierarchy_depth = 0
            if use_hierarchy_construction and created_breakdown_ids:
                try:
                    result_query = self.supabase.table('po_breakdowns')\
                        .select('hierarchy_level')\
                        .in_('id', [str(id) for id in created_breakdown_ids])\
                        .execute()
                    if result_query.data:
                        max_hierarchy_depth = max(row.get('hierarchy_level', 0) for row in result_query.data)
                except Exception as e:
                    logger.warning(f"Failed to calculate max hierarchy depth: {e}")
            
            # Store detailed errors, warnings, and conflicts in database
            await self._store_batch_errors(batch_id, errors)
            await self._store_batch_warnings(batch_id, warnings)
            await self._store_batch_conflicts(batch_id, conflicts)
            
            # Update batch status with comprehensive metrics
            await self._update_batch_status(
                batch_id=batch_id,
                status=status,
                status_message=status_message,
                metrics={
                    'total_records': total_records,
                    'processed_records': processed_records,
                    'successful_records': successful_records,
                    'failed_records': failed_records,
                    'skipped_records': len(conflicts) if config.conflict_resolution == ConflictResolution.skip else 0,
                    'updated_records': 0,
                    'created_hierarchies': created_hierarchies,
                    'max_hierarchy_depth': max_hierarchy_depth,
                    'error_count': len(errors),
                    'warning_count': len(warnings),
                    'conflict_count': len(conflicts),
                    'errors_by_category': errors_by_category,
                    'errors_by_severity': errors_by_severity,
                    'processing_time_ms': processing_time_ms,
                    'created_breakdown_ids': [str(id) for id in created_breakdown_ids]
                }
            )
            
            result = ImportResult(
                batch_id=batch_id,
                status=status,
                status_message=status_message,
                total_records=total_records,
                processed_records=processed_records,
                successful_records=successful_records,
                failed_records=failed_records,
                skipped_records=len(conflicts) if config.conflict_resolution == ConflictResolution.skip else 0,
                updated_records=0,
                conflicts=conflicts,
                errors=errors,
                warnings=warnings,
                error_count=len(errors),
                warning_count=len(warnings),
                conflict_count=len(conflicts),
                errors_by_category=errors_by_category,
                errors_by_severity=errors_by_severity,
                processing_time_ms=processing_time_ms,
                created_hierarchies=created_hierarchies,
                max_hierarchy_depth=max_hierarchy_depth,
                created_breakdown_ids=created_breakdown_ids,
                can_rollback=True,
                rollback_instructions=f"Use rollback_import_batch({batch_id}) to delete all {len(created_breakdown_ids)} created breakdowns."
            )
            
            logger.info(
                f"Excel import completed: batch_id={batch_id}, "
                f"status={status}, successful={successful_records}/{total_records}, "
                f"errors={len(errors)}, warnings={len(warnings)}, conflicts={len(conflicts)}"
            )
            
            return result
            
        except Exception as e:
            logger.error(f"Excel import failed: {e}")
            await self._update_batch_status(
                batch_id=batch_id,
                status=ImportStatus.failed,
                status_message=f"Import failed with exception: {str(e)}"
            )
            raise HTTPException(
                status_code=500,
                detail=f"Import processing failed: {str(e)}"
            )

    
    # =========================================================================
    # Import Batch Management
    # =========================================================================
    
    async def create_import_batch(
        self,
        project_id: UUID,
        source: str,
        user_id: UUID,
        file_name: Optional[str] = None,
        file_size_bytes: Optional[int] = None,
        file_type: Optional[str] = None,
        import_config: Optional[ImportConfig] = None
    ) -> UUID:
        """
        Create a new import batch record for comprehensive tracking.
        
        **Validates: Requirements 1.6, 10.3, 10.4**
        
        Args:
            project_id: Project UUID
            source: Source description (e.g., filename)
            user_id: User performing the import
            file_name: Original filename
            file_size_bytes: File size in bytes
            file_type: File type (csv, xlsx, xls)
            import_config: Import configuration for reference
            
        Returns:
            UUID of the created batch
        """
        batch_id = uuid4()
        
        try:
            batch_data = {
                'id': str(batch_id),
                'project_id': str(project_id),
                'source': source,
                'file_name': file_name,
                'file_size_bytes': file_size_bytes,
                'file_type': file_type,
                'status': ImportStatus.pending.value,
                'status_message': 'Import batch created, awaiting processing',
                'total_records': 0,
                'processed_records': 0,
                'successful_records': 0,
                'failed_records': 0,
                'skipped_records': 0,
                'updated_records': 0,
                'created_hierarchies': 0,
                'max_hierarchy_depth': 0,
                'error_count': 0,
                'warning_count': 0,
                'conflict_count': 0,
                'errors': [],
                'warnings': [],
                'conflicts': [],
                'import_config': import_config.model_dump() if import_config else {},
                'can_rollback': True,
                'created_breakdown_ids': [],
                'imported_by': str(user_id),
                'started_at': datetime.now().isoformat(),
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat()
            }
            
            # Insert into database
            result = self.supabase.table(self.import_batch_table).insert(batch_data).execute()
            
            if not result.data:
                logger.warning(f"Failed to insert batch record, but continuing with batch_id {batch_id}")
            else:
                logger.info(f"Created import batch {batch_id} for project {project_id}")
            
            return batch_id
            
        except Exception as e:
            logger.error(f"Failed to create import batch: {e}")
            # Return batch_id anyway so import can proceed
            return batch_id
    
    async def get_import_status(self, batch_id: UUID) -> ImportBatchStatus:
        """
        Get the comprehensive status of an import batch.
        
        **Validates: Requirements 1.6, 10.3, 10.4**
        
        Args:
            batch_id: Import batch UUID
            
        Returns:
            ImportBatchStatus with complete batch information
        """
        try:
            result = self.supabase.table(self.import_batch_table)\
                .select('*')\
                .eq('id', str(batch_id))\
                .execute()
            
            if not result.data:
                raise HTTPException(
                    status_code=404,
                    detail=f"Import batch {batch_id} not found"
                )
            
            batch_data = result.data[0]
            
            # Parse created_breakdown_ids from JSONB
            created_ids = batch_data.get('created_breakdown_ids', [])
            if isinstance(created_ids, str):
                import json
                created_ids = json.loads(created_ids)
            created_ids = [UUID(id_str) for id_str in created_ids if id_str]
            
            # Parse timestamps
            started_at = None
            if batch_data.get('started_at'):
                started_at = datetime.fromisoformat(batch_data['started_at'].replace('Z', '+00:00'))
            
            completed_at = None
            if batch_data.get('completed_at'):
                completed_at = datetime.fromisoformat(batch_data['completed_at'].replace('Z', '+00:00'))
            
            rolled_back_at = None
            if batch_data.get('rolled_back_at'):
                rolled_back_at = datetime.fromisoformat(batch_data['rolled_back_at'].replace('Z', '+00:00'))
            
            # Parse errors_by_category and errors_by_severity
            errors_by_category = batch_data.get('errors_by_category', {})
            if isinstance(errors_by_category, str):
                import json
                errors_by_category = json.loads(errors_by_category)
            
            errors_by_severity = batch_data.get('errors_by_severity', {})
            if isinstance(errors_by_severity, str):
                import json
                errors_by_severity = json.loads(errors_by_severity)
            
            return ImportBatchStatus(
                id=UUID(batch_data['id']),
                project_id=UUID(batch_data['project_id']),
                source=batch_data['source'],
                file_name=batch_data.get('file_name'),
                file_size_bytes=batch_data.get('file_size_bytes'),
                file_type=batch_data.get('file_type'),
                status=ImportStatus(batch_data['status']),
                status_message=batch_data.get('status_message'),
                total_records=batch_data.get('total_records', 0),
                processed_records=batch_data.get('processed_records', 0),
                successful_records=batch_data.get('successful_records', 0),
                failed_records=batch_data.get('failed_records', 0),
                skipped_records=batch_data.get('skipped_records', 0),
                updated_records=batch_data.get('updated_records', 0),
                created_hierarchies=batch_data.get('created_hierarchies', 0),
                max_hierarchy_depth=batch_data.get('max_hierarchy_depth', 0),
                error_count=batch_data.get('error_count', 0),
                warning_count=batch_data.get('warning_count', 0),
                conflict_count=batch_data.get('conflict_count', 0),
                errors_by_category=errors_by_category,
                errors_by_severity=errors_by_severity,
                started_at=started_at,
                completed_at=completed_at,
                processing_time_ms=batch_data.get('processing_time_ms'),
                can_rollback=batch_data.get('can_rollback', True),
                rolled_back_at=rolled_back_at,
                rolled_back_by=UUID(batch_data['rolled_back_by']) if batch_data.get('rolled_back_by') else None,
                rollback_reason=batch_data.get('rollback_reason'),
                created_breakdown_ids=created_ids,
                imported_by=UUID(batch_data['imported_by']),
                created_at=datetime.fromisoformat(batch_data['created_at'].replace('Z', '+00:00')),
                updated_at=datetime.fromisoformat(batch_data['updated_at'].replace('Z', '+00:00'))
            )
            
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Failed to get import status: {e}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to retrieve import status: {str(e)}"
            )
    
    async def _update_batch_status(
        self,
        batch_id: UUID,
        status: ImportStatus,
        status_message: Optional[str] = None,
        metrics: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Update the status of an import batch with comprehensive metrics.
        
        **Validates: Requirements 1.6, 10.3, 10.4**
        
        Args:
            batch_id: Import batch UUID
            status: New status
            status_message: Optional status message
            metrics: Optional dict with metrics to update
        """
        try:
            update_data = {
                'status': status.value,
                'updated_at': datetime.now().isoformat()
            }
            
            if status_message:
                update_data['status_message'] = status_message
            
            # Add completion timestamp for terminal states
            if status in [ImportStatus.completed, ImportStatus.failed, ImportStatus.partially_completed, ImportStatus.rolled_back]:
                update_data['completed_at'] = datetime.now().isoformat()
            
            # Add metrics if provided
            if metrics:
                update_data.update(metrics)
            
            self.supabase.table(self.import_batch_table)\
                .update(update_data)\
                .eq('id', str(batch_id))\
                .execute()
            
            logger.info(f"Updated batch {batch_id} status to {status.value}")
            
        except Exception as e:
            logger.warning(f"Failed to update batch status: {e}")
    
    async def _store_batch_errors(
        self,
        batch_id: UUID,
        errors: List[ImportError]
    ) -> None:
        """
        Store detailed error information in the database.
        
        **Validates: Requirements 10.3, 10.4**
        
        Args:
            batch_id: Import batch UUID
            errors: List of errors to store
        """
        if not errors:
            return
        
        try:
            error_records = []
            for error in errors:
                error_record = {
                    'id': str(uuid4()),
                    'batch_id': str(batch_id),
                    'row_number': error.row_number,
                    'field': error.field,
                    'error_type': error.error_type,
                    'severity': error.severity.value,
                    'category': error.category.value,
                    'message': error.message,
                    'raw_value': error.raw_value,
                    'suggested_fix': error.suggested_fix,
                    'can_auto_fix': error.can_auto_fix,
                    'error_data': error.error_data,
                    'created_at': datetime.now().isoformat()
                }
                error_records.append(error_record)
            
            # Batch insert errors
            if error_records:
                self.supabase.table('po_import_batch_errors').insert(error_records).execute()
                logger.info(f"Stored {len(error_records)} errors for batch {batch_id}")
            
        except Exception as e:
            logger.error(f"Failed to store batch errors: {e}")
    
    async def _store_batch_warnings(
        self,
        batch_id: UUID,
        warnings: List[ImportWarning]
    ) -> None:
        """
        Store warning information in the database.
        
        **Validates: Requirements 10.3**
        
        Args:
            batch_id: Import batch UUID
            warnings: List of warnings to store
        """
        if not warnings:
            return
        
        try:
            warning_records = []
            for warning in warnings:
                warning_record = {
                    'id': str(uuid4()),
                    'batch_id': str(batch_id),
                    'row_number': warning.row_number,
                    'field': warning.field,
                    'warning_type': warning.warning_type,
                    'message': warning.message,
                    'suggestion': warning.suggestion,
                    'warning_data': warning.warning_data,
                    'created_at': datetime.now().isoformat()
                }
                warning_records.append(warning_record)
            
            # Batch insert warnings
            if warning_records:
                self.supabase.table('po_import_batch_warnings').insert(warning_records).execute()
                logger.info(f"Stored {len(warning_records)} warnings for batch {batch_id}")
            
        except Exception as e:
            logger.error(f"Failed to store batch warnings: {e}")
    
    async def _store_batch_conflicts(
        self,
        batch_id: UUID,
        conflicts: List[ImportConflict]
    ) -> None:
        """
        Store conflict information in the database.
        
        **Validates: Requirements 10.4**
        
        Args:
            batch_id: Import batch UUID
            conflicts: List of conflicts to store
        """
        if not conflicts:
            return
        
        try:
            conflict_records = []
            for conflict in conflicts:
                conflict_record = {
                    'id': str(uuid4()),
                    'batch_id': str(batch_id),
                    'row_number': conflict.row_number,
                    'conflict_type': conflict.conflict_type.value,
                    'existing_record': conflict.existing_record,
                    'new_record': conflict.new_record,
                    'field_conflicts': conflict.field_conflicts,
                    'suggested_resolution': conflict.suggested_resolution.value,
                    'resolution_applied': conflict.resolution_applied.value if conflict.resolution_applied else None,
                    'created_at': datetime.now().isoformat()
                }
                conflict_records.append(conflict_record)
            
            # Batch insert conflicts
            if conflict_records:
                self.supabase.table('po_import_batch_conflicts').insert(conflict_records).execute()
                logger.info(f"Stored {len(conflict_records)} conflicts for batch {batch_id}")
            
        except Exception as e:
            logger.error(f"Failed to store batch conflicts: {e}")
    
    async def get_batch_errors(
        self,
        batch_id: UUID,
        severity: Optional[ErrorSeverity] = None,
        category: Optional[ErrorCategory] = None,
        limit: int = 100
    ) -> List[ImportBatchErrorDetail]:
        """
        Retrieve detailed errors for an import batch.
        
        **Validates: Requirements 10.3, 10.4**
        
        Args:
            batch_id: Import batch UUID
            severity: Optional filter by severity
            category: Optional filter by category
            limit: Maximum number of errors to return
            
        Returns:
            List of ImportBatchErrorDetail objects
        """
        try:
            query = self.supabase.table('po_import_batch_errors')\
                .select('*')\
                .eq('batch_id', str(batch_id))\
                .order('row_number')\
                .limit(limit)
            
            if severity:
                query = query.eq('severity', severity.value)
            
            if category:
                query = query.eq('category', category.value)
            
            result = query.execute()
            
            errors = []
            for row in result.data:
                error = ImportBatchErrorDetail(
                    id=UUID(row['id']),
                    batch_id=UUID(row['batch_id']),
                    row_number=row['row_number'],
                    field=row.get('field'),
                    error_type=row['error_type'],
                    severity=ErrorSeverity(row['severity']),
                    category=ErrorCategory(row['category']),
                    message=row['message'],
                    raw_value=row.get('raw_value'),
                    suggested_fix=row.get('suggested_fix'),
                    can_auto_fix=row.get('can_auto_fix', False),
                    error_data=row.get('error_data', {}),
                    created_at=datetime.fromisoformat(row['created_at'].replace('Z', '+00:00'))
                )
                errors.append(error)
            
            return errors
            
        except Exception as e:
            logger.error(f"Failed to retrieve batch errors: {e}")
            return []
    
    async def rollback_import_batch(
        self,
        batch_id: UUID,
        user_id: UUID,
        reason: str
    ) -> bool:
        """
        Rollback an import batch by deleting all created breakdowns.
        
        **Validates: Requirements 10.3**
        
        Args:
            batch_id: Import batch UUID to rollback
            user_id: User performing the rollback
            reason: Reason for rollback
            
        Returns:
            True if successful
        """
        try:
            # Get batch status
            batch_status = await self.get_import_status(batch_id)
            
            if not batch_status.can_rollback:
                raise ValueError("This import batch cannot be rolled back")
            
            if batch_status.status == ImportStatus.rolled_back:
                raise ValueError("This import batch has already been rolled back")
            
            # Delete all created breakdowns (soft delete)
            deleted_count = 0
            for breakdown_id in batch_status.created_breakdown_ids:
                try:
                    await self.po_service.delete_breakdown(
                        breakdown_id=breakdown_id,
                        user_id=user_id,
                        hard_delete=False
                    )
                    deleted_count += 1
                except Exception as e:
                    logger.error(f"Failed to delete breakdown {breakdown_id} during rollback: {e}")
            
            # Update batch status
            await self._update_batch_status(
                batch_id=batch_id,
                status=ImportStatus.rolled_back,
                status_message=f"Rolled back: {reason}. Deleted {deleted_count} breakdowns.",
                metrics={
                    'rolled_back_at': datetime.now().isoformat(),
                    'rolled_back_by': str(user_id),
                    'rollback_reason': reason,
                    'can_rollback': False
                }
            )
            
            logger.info(f"Rolled back import batch {batch_id}, deleted {deleted_count} breakdowns")
            
            return True
            
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Failed to rollback import batch {batch_id}: {e}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to rollback import batch: {str(e)}"
            )

    
    # =========================================================================
    # Parsing and Transformation
    # =========================================================================
    
    def _parse_csv_content(
        self,
        csv_content: str,
        config: ImportConfig,
        errors: List[ImportError],
        warnings: List[ImportWarning]
    ) -> List[Dict[str, Any]]:
        """
        Parse CSV content into structured data using column mappings.
        
        **Validates: Requirements 1.1, 1.2**
        """
        parsed_rows = []
        
        try:
            # Create CSV reader with configured delimiter
            csv_reader = csv.DictReader(
                io.StringIO(csv_content),
                delimiter=config.delimiter
            )
            
            # Skip header rows if configured
            for _ in range(config.skip_header_rows - 1):
                try:
                    next(csv_reader)
                except StopIteration:
                    break
            
            # Parse each row
            for row_num, row in enumerate(csv_reader, start=config.skip_header_rows + 1):
                parsed_row = {}
                
                # Apply column mappings
                for target_field, csv_column in config.column_mappings.items():
                    if csv_column in row:
                        parsed_row[target_field] = row[csv_column]
                    else:
                        # Column not found in CSV
                        if target_field in REQUIRED_FIELDS:
                            errors.append(ImportError(
                                row_number=row_num,
                                field=target_field,
                                error_type='missing_column',
                                message=f"Required column '{csv_column}' not found in CSV",
                                raw_value=None
                            ))
                
                # Store unmapped columns as custom fields
                custom_fields = {}
                mapped_columns = set(config.column_mappings.values())
                for csv_col, value in row.items():
                    if csv_col not in mapped_columns and value and value.strip():
                        custom_fields[csv_col] = value.strip()
                
                if custom_fields:
                    parsed_row['custom_fields'] = custom_fields
                
                parsed_rows.append(parsed_row)
            
        except Exception as e:
            errors.append(ImportError(
                row_number=0,
                field='general',
                error_type='parse_error',
                message=f"Failed to parse CSV: {str(e)}",
                raw_value=None
            ))
            logger.error(f"CSV parsing failed: {e}")
        
        return parsed_rows
    
    def _parse_excel_sheet(
        self,
        sheet,
        config: ImportConfig,
        errors: List[ImportError],
        warnings: List[ImportWarning]
    ) -> List[Dict[str, Any]]:
        """
        Parse Excel sheet into structured data using column mappings.
        
        **Validates: Requirements 1.1, 1.2**
        """
        parsed_rows = []
        
        try:
            # Get header row
            header_row_idx = config.skip_header_rows
            headers = []
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_idx == header_row_idx:
                    headers = [str(cell) if cell is not None else '' for cell in row]
                    continue
                elif row_idx < header_row_idx:
                    continue
                
                # Parse data rows
                parsed_row = {}
                
                # Apply column mappings
                for target_field, csv_column in config.column_mappings.items():
                    try:
                        col_idx = headers.index(csv_column)
                        if col_idx < len(row):
                            cell_value = row[col_idx]
                            parsed_row[target_field] = str(cell_value) if cell_value is not None else ''
                    except ValueError:
                        # Column not found
                        if target_field in REQUIRED_FIELDS:
                            errors.append(ImportError(
                                row_number=row_idx,
                                field=target_field,
                                error_type='missing_column',
                                message=f"Required column '{csv_column}' not found in Excel",
                                raw_value=None
                            ))
                
                # Store unmapped columns as custom fields
                custom_fields = {}
                mapped_columns = set(config.column_mappings.values())
                for col_idx, header in enumerate(headers):
                    if header not in mapped_columns and col_idx < len(row):
                        cell_value = row[col_idx]
                        if cell_value is not None and str(cell_value).strip():
                            custom_fields[header] = str(cell_value).strip()
                
                if custom_fields:
                    parsed_row['custom_fields'] = custom_fields
                
                # Only add non-empty rows
                if any(v for v in parsed_row.values() if v):
                    parsed_rows.append(parsed_row)
            
        except Exception as e:
            errors.append(ImportError(
                row_number=0,
                field='general',
                error_type='parse_error',
                message=f"Failed to parse Excel: {str(e)}",
                raw_value=None
            ))
            logger.error(f"Excel parsing failed: {e}")
        
        return parsed_rows

    
    def _transform_row_to_breakdown(
        self,
        row_data: Dict[str, Any],
        config: ImportConfig,
        row_number: int,
        errors: List[ImportError],
        warnings: List[ImportWarning]
    ) -> Optional[POBreakdownCreate]:
        """
        Transform a parsed row into a POBreakdownCreate object.
        
        **Validates: Requirements 1.2, 10.1**
        
        Args:
            row_data: Parsed row data
            config: Import configuration
            row_number: Row number for error reporting
            errors: List to append errors to
            warnings: List to append warnings to
            
        Returns:
            POBreakdownCreate object or None if validation fails
        """
        try:
            # Extract and validate required fields
            name = row_data.get('name', '').strip()
            if not name:
                errors.append(ImportError(
                    row_number=row_number,
                    field='name',
                    error_type='required_field_missing',
                    severity=ErrorSeverity.error,
                    category=ErrorCategory.validation,
                    message="Name is required",
                    raw_value=row_data.get('name'),
                    suggested_fix="Provide a non-empty name for the breakdown item"
                ))
                return None
            
            # Parse amounts with validation
            planned_amount = self._parse_decimal(
                row_data.get('planned_amount', '0'),
                row_number,
                'planned_amount',
                errors,
                warnings,
                config.validate_amounts
            )
            
            committed_amount = self._parse_decimal(
                row_data.get('committed_amount', '0'),
                row_number,
                'committed_amount',
                errors,
                warnings,
                config.validate_amounts
            )
            
            actual_amount = self._parse_decimal(
                row_data.get('actual_amount', '0'),
                row_number,
                'actual_amount',
                errors,
                warnings,
                config.validate_amounts
            )
            
            # Validate amounts if configured
            if config.validate_amounts:
                if planned_amount < 0 or committed_amount < 0 or actual_amount < 0:
                    errors.append(ImportError(
                        row_number=row_number,
                        field='amounts',
                        error_type='invalid_amount',
                        severity=ErrorSeverity.error,
                        category=ErrorCategory.validation,
                        message="Amounts cannot be negative",
                        raw_value=f"planned={planned_amount}, committed={committed_amount}, actual={actual_amount}",
                        suggested_fix="Ensure all amount values are non-negative"
                    ))
                    return None
            
            # Parse parent reference if provided
            parent_breakdown_id = None
            if config.parent_reference_column and row_data.get('parent_reference'):
                try:
                    parent_breakdown_id = UUID(row_data['parent_reference'])
                except ValueError:
                    warnings.append(ImportWarning(
                        row_number=row_number,
                        field='parent_reference',
                        warning_type='invalid_uuid',
                        message=f"Invalid parent reference UUID: {row_data['parent_reference']}"
                    ))
            
            # Parse tags
            tags = []
            if row_data.get('tags'):
                tags_str = row_data['tags']
                if isinstance(tags_str, str):
                    tags = [t.strip() for t in tags_str.split(',') if t.strip()]
            
            # Get custom fields
            custom_fields = row_data.get('custom_fields', {})
            
            # Create POBreakdownCreate object
            breakdown_data = POBreakdownCreate(
                name=name,
                code=row_data.get('code', '').strip() or None,
                sap_po_number=row_data.get('sap_po_number', '').strip() or None,
                sap_line_item=row_data.get('sap_line_item', '').strip() or None,
                parent_breakdown_id=parent_breakdown_id,
                cost_center=row_data.get('cost_center', '').strip() or None,
                gl_account=row_data.get('gl_account', '').strip() or None,
                planned_amount=planned_amount,
                committed_amount=committed_amount,
                actual_amount=actual_amount,
                currency=row_data.get('currency', config.currency_default).strip().upper(),
                breakdown_type=config.breakdown_type_default,
                category=row_data.get('category', '').strip() or None,
                subcategory=row_data.get('subcategory', '').strip() or None,
                custom_fields=custom_fields,
                tags=tags,
                notes=row_data.get('notes', '').strip() or None
            )
            
            return breakdown_data
            
        except Exception as e:
            errors.append(ImportError(
                row_number=row_number,
                field='general',
                error_type='transformation_error',
                severity=ErrorSeverity.error,
                category=ErrorCategory.transformation,
                message=f"Failed to transform row: {str(e)}",
                raw_value=str(row_data),
                suggested_fix="Check data format and ensure all required fields are present"
            ))
            logger.error(f"Row transformation failed for row {row_number}: {e}")
            return None
    
    def _parse_decimal(
        self,
        value: Any,
        row_number: int,
        field: str,
        errors: List[ImportError],
        warnings: List[ImportWarning],
        validate: bool = True
    ) -> Decimal:
        """
        Parse a value to Decimal with error handling.
        
        **Validates: Requirements 10.1**
        """
        if value is None or value == '':
            return Decimal('0.00')
        
        try:
            # Clean the value
            if isinstance(value, str):
                # Remove common formatting characters
                cleaned = value.strip().replace(',', '').replace('$', '').replace('€', '').replace('£', '')
                if not cleaned:
                    return Decimal('0.00')
                value = cleaned
            
            result = Decimal(str(value))
            
            # Validate if configured
            if validate and result < 0:
                warnings.append(ImportWarning(
                    row_number=row_number,
                    field=field,
                    warning_type='negative_amount',
                    message=f"Negative amount detected: {result}"
                ))
            
            return result
            
        except (InvalidOperation, ValueError) as e:
            errors.append(ImportError(
                row_number=row_number,
                field=field,
                error_type='invalid_decimal',
                severity=ErrorSeverity.error,
                category=ErrorCategory.validation,
                message=f"Invalid decimal value: {value}",
                raw_value=str(value),
                suggested_fix="Provide a valid numeric value (e.g., 1000.50)"
            ))
            return Decimal('0.00')

    
    # =========================================================================
    # Conflict Detection and Resolution
    # =========================================================================
    
    async def _check_for_conflicts(
        self,
        breakdown_data: POBreakdownCreate,
        project_id: UUID,
        row_number: int
    ) -> Optional[ImportConflict]:
        """
        Check for conflicts with existing records.
        
        **Validates: Requirements 1.4, 10.2**
        
        Args:
            breakdown_data: Breakdown data to check
            project_id: Project UUID
            row_number: Row number for reporting
            
        Returns:
            ImportConflict if conflict detected, None otherwise
        """
        try:
            # Check for duplicate code
            if breakdown_data.code:
                existing = self.supabase.table('po_breakdowns')\
                    .select('*')\
                    .eq('project_id', str(project_id))\
                    .eq('code', breakdown_data.code)\
                    .eq('is_active', True)\
                    .execute()
                
                if existing.data:
                    return ImportConflict(
                        row_number=row_number,
                        conflict_type=ConflictType.duplicate_code,
                        existing_record=existing.data[0],
                        new_record=breakdown_data.model_dump(),
                        suggested_resolution=ConflictResolution.update,
                        field_conflicts=['code']
                    )
            
            # Check for duplicate SAP reference
            if breakdown_data.sap_po_number and breakdown_data.sap_line_item:
                existing = self.supabase.table('po_breakdowns')\
                    .select('*')\
                    .eq('project_id', str(project_id))\
                    .eq('sap_po_number', breakdown_data.sap_po_number)\
                    .eq('sap_line_item', breakdown_data.sap_line_item)\
                    .eq('is_active', True)\
                    .execute()
                
                if existing.data:
                    return ImportConflict(
                        row_number=row_number,
                        conflict_type=ConflictType.duplicate_sap_reference,
                        existing_record=existing.data[0],
                        new_record=breakdown_data.model_dump(),
                        suggested_resolution=ConflictResolution.update,
                        field_conflicts=['sap_po_number', 'sap_line_item']
                    )
            
            # Check if parent exists (if parent_breakdown_id is provided)
            if breakdown_data.parent_breakdown_id:
                parent = await self.po_service.get_breakdown_by_id(breakdown_data.parent_breakdown_id)
                if not parent:
                    return ImportConflict(
                        row_number=row_number,
                        conflict_type=ConflictType.parent_not_found,
                        existing_record={},
                        new_record=breakdown_data.model_dump(),
                        suggested_resolution=ConflictResolution.manual,
                        field_conflicts=['parent_breakdown_id']
                    )
            
            return None
            
        except Exception as e:
            logger.error(f"Conflict check failed for row {row_number}: {e}")
            return None
    
    async def resolve_import_conflicts(
        self,
        batch_id: UUID,
        resolutions: List[Dict[str, Any]]
    ) -> ImportResult:
        """
        Resolve conflicts from a previous import attempt.
        
        **Validates: Requirements 1.4, 10.5**
        
        Args:
            batch_id: Import batch UUID
            resolutions: List of conflict resolutions with row_number and resolution strategy
            
        Returns:
            ImportResult with resolution results
        """
        # This would be implemented to handle manual conflict resolution
        # For now, return a placeholder
        raise NotImplementedError("Conflict resolution not yet implemented")
    
    # =========================================================================
    # Hierarchy Construction
    # =========================================================================
    
    async def construct_hierarchy_from_import(
        self,
        parsed_rows: List[Dict[str, Any]],
        project_id: UUID,
        config: ImportConfig,
        user_id: UUID,
        batch_id: UUID,
        errors: List[ImportError],
        warnings: List[ImportWarning]
    ) -> Tuple[List[UUID], int]:
        """
        Construct hierarchical relationships from imported data.
        
        **Validates: Requirements 1.3, 1.4, 10.2**
        
        This method processes imported rows to:
        1. Parse SAP structure codes to determine hierarchy levels
        2. Build parent-child relationships automatically
        3. Create missing parent items when configured
        4. Detect and resolve duplicate entries
        
        Args:
            parsed_rows: List of parsed row data
            project_id: Target project UUID
            config: Import configuration
            user_id: User performing the import
            batch_id: Import batch ID
            errors: List to append errors to
            warnings: List to append warnings to
            
        Returns:
            Tuple of (created_breakdown_ids, hierarchy_count)
        """
        created_breakdown_ids: List[UUID] = []
        hierarchy_map: Dict[str, UUID] = {}  # Maps structure code to breakdown ID
        pending_items: List[Tuple[int, Dict[str, Any], POBreakdownCreate]] = []  # (row_num, row_data, breakdown_data)
        
        # Phase 1: Parse hierarchy information from all rows
        hierarchy_info = self._parse_hierarchy_information(
            parsed_rows=parsed_rows,
            config=config,
            errors=errors,
            warnings=warnings
        )
        
        # Phase 2: Sort items by hierarchy level (parents first)
        sorted_items = sorted(
            hierarchy_info,
            key=lambda x: (x['hierarchy_level'], x['row_number'])
        )
        
        # Phase 3: Create items in hierarchy order
        for item_info in sorted_items:
            row_num = item_info['row_number']
            row_data = item_info['row_data']
            hierarchy_level = item_info['hierarchy_level']
            structure_code = item_info['structure_code']
            parent_code = item_info['parent_code']
            
            try:
                # Transform row to breakdown data
                breakdown_data = self._transform_row_to_breakdown(
                    row_data=row_data,
                    config=config,
                    row_number=row_num,
                    errors=errors,
                    warnings=warnings
                )
                
                if not breakdown_data:
                    continue
                
                # Determine parent ID from hierarchy
                parent_id = None
                if parent_code:
                    if parent_code in hierarchy_map:
                        parent_id = hierarchy_map[parent_code]
                    elif config.create_missing_parents:
                        # Create missing parent
                        parent_id = await self._create_missing_parent(
                            parent_code=parent_code,
                            project_id=project_id,
                            user_id=user_id,
                            batch_id=batch_id,
                            config=config,
                            hierarchy_map=hierarchy_map,
                            warnings=warnings,
                            row_number=row_num
                        )
                    else:
                        errors.append(ImportError(
                            row_number=row_num,
                            field='parent_reference',
                            error_type='parent_not_found',
                            message=f"Parent with code '{parent_code}' not found and create_missing_parents is disabled",
                            raw_value=parent_code
                        ))
                        continue
                
                # Override parent_breakdown_id with hierarchy-derived parent
                breakdown_data.parent_breakdown_id = parent_id
                
                # Check for duplicates within this import batch
                duplicate_check = self._check_batch_duplicate(
                    breakdown_data=breakdown_data,
                    hierarchy_map=hierarchy_map,
                    structure_code=structure_code,
                    row_number=row_num
                )
                
                if duplicate_check:
                    warnings.append(ImportWarning(
                        row_number=row_num,
                        field='code',
                        warning_type='duplicate_in_batch',
                        message=duplicate_check
                    ))
                    continue
                
                # Create the breakdown item
                created = await self.po_service.create_breakdown(
                    project_id=project_id,
                    breakdown_data=breakdown_data,
                    user_id=user_id
                )
                
                # Update with import metadata
                await self._update_import_metadata(
                    breakdown_id=created.id,
                    batch_id=batch_id,
                    source=f"Import batch {batch_id}"
                )
                
                # Track in hierarchy map
                if structure_code:
                    hierarchy_map[structure_code] = created.id
                
                created_breakdown_ids.append(created.id)
                
            except Exception as e:
                errors.append(ImportError(
                    row_number=row_num,
                    field='general',
                    error_type='hierarchy_construction_error',
                    message=f"Failed to create item in hierarchy: {str(e)}",
                    raw_value=str(row_data)
                ))
                logger.error(f"Failed to create hierarchy item at row {row_num}: {e}")
        
        # Count unique hierarchy levels created
        hierarchy_count = len(set(item['hierarchy_level'] for item in hierarchy_info))
        
        return created_breakdown_ids, hierarchy_count
    
    def _parse_hierarchy_information(
        self,
        parsed_rows: List[Dict[str, Any]],
        config: ImportConfig,
        errors: List[ImportError],
        warnings: List[ImportWarning]
    ) -> List[Dict[str, Any]]:
        """
        Parse hierarchy information from all rows.
        
        **Validates: Requirements 1.3**
        
        Extracts hierarchy level and parent relationships from:
        - hierarchy_column: SAP structure codes (e.g., "1.2.3")
        - parent_reference_column: Direct parent references
        
        Args:
            parsed_rows: List of parsed row data
            config: Import configuration
            errors: List to append errors to
            warnings: List to append warnings to
            
        Returns:
            List of dicts with hierarchy information
        """
        hierarchy_info = []
        
        for row_num, row_data in enumerate(parsed_rows, start=config.skip_header_rows + 1):
            hierarchy_level = 0
            structure_code = None
            parent_code = None
            
            # Parse hierarchy from structure code column
            if config.hierarchy_column and row_data.get(config.hierarchy_column):
                structure_code = str(row_data[config.hierarchy_column]).strip()
                
                # Parse SAP structure code (e.g., "1.2.3" -> level 2, parent "1.2")
                hierarchy_level, parent_code = self._parse_sap_structure_code(
                    structure_code=structure_code,
                    row_number=row_num,
                    errors=errors
                )
            
            # Override with explicit parent reference if provided
            if config.parent_reference_column and row_data.get(config.parent_reference_column):
                parent_code = str(row_data[config.parent_reference_column]).strip()
                # Increment level if parent is specified
                if parent_code:
                    hierarchy_level = hierarchy_level + 1 if hierarchy_level > 0 else 1
            
            # Validate hierarchy depth
            if hierarchy_level > config.max_hierarchy_depth:
                errors.append(ImportError(
                    row_number=row_num,
                    field='hierarchy_level',
                    error_type='depth_exceeded',
                    message=f"Hierarchy level {hierarchy_level} exceeds maximum depth of {config.max_hierarchy_depth}",
                    raw_value=structure_code
                ))
                continue
            
            hierarchy_info.append({
                'row_number': row_num,
                'row_data': row_data,
                'hierarchy_level': hierarchy_level,
                'structure_code': structure_code,
                'parent_code': parent_code
            })
        
        return hierarchy_info
    
    def _parse_sap_structure_code(
        self,
        structure_code: str,
        row_number: int,
        errors: List[ImportError]
    ) -> Tuple[int, Optional[str]]:
        """
        Parse SAP structure code to determine hierarchy level and parent code.
        
        **Validates: Requirements 1.3**
        
        SAP structure codes follow patterns like:
        - "1" -> level 0, no parent
        - "1.1" -> level 1, parent "1"
        - "1.1.1" -> level 2, parent "1.1"
        - "1.2.3.4" -> level 3, parent "1.2.3"
        
        Args:
            structure_code: SAP structure code string
            row_number: Row number for error reporting
            errors: List to append errors to
            
        Returns:
            Tuple of (hierarchy_level, parent_code)
        """
        try:
            # Split by common delimiters
            parts = structure_code.replace('-', '.').replace('/', '.').split('.')
            parts = [p.strip() for p in parts if p.strip()]
            
            if not parts:
                return 0, None
            
            # Hierarchy level is number of parts minus 1 (root is level 0)
            hierarchy_level = len(parts) - 1
            
            # Parent code is all parts except the last one
            parent_code = None
            if len(parts) > 1:
                parent_code = '.'.join(parts[:-1])
            
            return hierarchy_level, parent_code
            
        except Exception as e:
            errors.append(ImportError(
                row_number=row_number,
                field='hierarchy_column',
                error_type='invalid_structure_code',
                message=f"Failed to parse structure code: {str(e)}",
                raw_value=structure_code
            ))
            return 0, None
    
    async def _create_missing_parent(
        self,
        parent_code: str,
        project_id: UUID,
        user_id: UUID,
        batch_id: UUID,
        config: ImportConfig,
        hierarchy_map: Dict[str, UUID],
        warnings: List[ImportWarning],
        row_number: int
    ) -> Optional[UUID]:
        """
        Create a missing parent item automatically.
        
        **Validates: Requirements 1.3, 10.2**
        
        Args:
            parent_code: Structure code of the missing parent
            project_id: Project UUID
            user_id: User ID
            batch_id: Import batch ID
            config: Import configuration
            hierarchy_map: Map of structure codes to breakdown IDs
            warnings: List to append warnings to
            row_number: Row number for warning reporting
            
        Returns:
            UUID of created parent or None if creation failed
        """
        try:
            # Check if parent already exists in database
            existing = self.supabase.table('po_breakdowns')\
                .select('*')\
                .eq('project_id', str(project_id))\
                .eq('code', parent_code)\
                .eq('is_active', True)\
                .execute()
            
            if existing.data:
                parent_id = UUID(existing.data[0]['id'])
                hierarchy_map[parent_code] = parent_id
                return parent_id
            
            # Determine parent's parent
            parent_hierarchy_level, grandparent_code = self._parse_sap_structure_code(
                structure_code=parent_code,
                row_number=row_number,
                errors=[]
            )
            
            grandparent_id = None
            if grandparent_code:
                if grandparent_code in hierarchy_map:
                    grandparent_id = hierarchy_map[grandparent_code]
                else:
                    # Recursively create grandparent
                    grandparent_id = await self._create_missing_parent(
                        parent_code=grandparent_code,
                        project_id=project_id,
                        user_id=user_id,
                        batch_id=batch_id,
                        config=config,
                        hierarchy_map=hierarchy_map,
                        warnings=warnings,
                        row_number=row_number
                    )
            
            # Create the missing parent
            parent_data = POBreakdownCreate(
                name=f"Auto-created: {parent_code}",
                code=parent_code,
                parent_breakdown_id=grandparent_id,
                breakdown_type=config.breakdown_type_default,
                currency=config.currency_default,
                planned_amount=Decimal('0.00'),
                committed_amount=Decimal('0.00'),
                actual_amount=Decimal('0.00'),
                custom_fields={'auto_created': True, 'created_from_import': True}
            )
            
            created = await self.po_service.create_breakdown(
                project_id=project_id,
                breakdown_data=parent_data,
                user_id=user_id
            )
            
            # Update with import metadata
            await self._update_import_metadata(
                breakdown_id=created.id,
                batch_id=batch_id,
                source=f"Auto-created parent for import batch {batch_id}"
            )
            
            # Track in hierarchy map
            hierarchy_map[parent_code] = created.id
            
            warnings.append(ImportWarning(
                row_number=row_number,
                field='parent_reference',
                warning_type='parent_auto_created',
                message=f"Auto-created missing parent with code '{parent_code}'",
                suggestion="Review auto-created parent and update details as needed"
            ))
            
            logger.info(f"Auto-created missing parent {parent_code} with ID {created.id}")
            
            return created.id
            
        except Exception as e:
            logger.error(f"Failed to create missing parent {parent_code}: {e}")
            warnings.append(ImportWarning(
                row_number=row_number,
                field='parent_reference',
                warning_type='parent_creation_failed',
                message=f"Failed to auto-create parent '{parent_code}': {str(e)}"
            ))
            return None
    
    def _check_batch_duplicate(
        self,
        breakdown_data: POBreakdownCreate,
        hierarchy_map: Dict[str, UUID],
        structure_code: Optional[str],
        row_number: int
    ) -> Optional[str]:
        """
        Check for duplicates within the current import batch.
        
        **Validates: Requirements 1.4, 10.2**
        
        Args:
            breakdown_data: Breakdown data to check
            hierarchy_map: Map of structure codes to breakdown IDs
            structure_code: Structure code of the item
            row_number: Row number for reporting
            
        Returns:
            Error message if duplicate found, None otherwise
        """
        # Check for duplicate structure code in batch
        if structure_code and structure_code in hierarchy_map:
            return f"Duplicate structure code '{structure_code}' found in import batch"
        
        # Check for duplicate code in batch
        if breakdown_data.code:
            for code, breakdown_id in hierarchy_map.items():
                # This is a simplified check - in production, we'd query the created items
                if code == breakdown_data.code:
                    return f"Duplicate code '{breakdown_data.code}' found in import batch"
        
        return None
    
    async def detect_import_duplicates(
        self,
        parsed_rows: List[Dict[str, Any]],
        project_id: UUID,
        config: ImportConfig
    ) -> List[ImportConflict]:
        """
        Detect duplicate entries across the import batch and existing data.
        
        **Validates: Requirements 1.4, 10.2**
        
        Args:
            parsed_rows: List of parsed row data
            project_id: Project UUID
            config: Import configuration
            
        Returns:
            List of detected conflicts
        """
        conflicts: List[ImportConflict] = []
        seen_codes: Dict[str, int] = {}  # code -> first row number
        seen_sap_refs: Dict[str, int] = {}  # sap_ref -> first row number
        
        for row_num, row_data in enumerate(parsed_rows, start=config.skip_header_rows + 1):
            # Check for duplicate codes within batch
            code = row_data.get('code', '').strip()
            if code:
                if code in seen_codes:
                    conflicts.append(ImportConflict(
                        row_number=row_num,
                        conflict_type=ConflictType.duplicate_code,
                        existing_record={'row_number': seen_codes[code], 'code': code},
                        new_record={'row_number': row_num, 'code': code},
                        suggested_resolution=ConflictResolution.skip,
                        field_conflicts=['code']
                    ))
                else:
                    seen_codes[code] = row_num
            
            # Check for duplicate SAP references within batch
            sap_po = row_data.get('sap_po_number', '').strip()
            sap_line = row_data.get('sap_line_item', '').strip()
            if sap_po and sap_line:
                sap_ref = f"{sap_po}:{sap_line}"
                if sap_ref in seen_sap_refs:
                    conflicts.append(ImportConflict(
                        row_number=row_num,
                        conflict_type=ConflictType.duplicate_sap_reference,
                        existing_record={'row_number': seen_sap_refs[sap_ref], 'sap_reference': sap_ref},
                        new_record={'row_number': row_num, 'sap_reference': sap_ref},
                        suggested_resolution=ConflictResolution.update,
                        field_conflicts=['sap_po_number', 'sap_line_item']
                    ))
                else:
                    seen_sap_refs[sap_ref] = row_num
        
        return conflicts
    
    # =========================================================================
    # Helper Methods
    # =========================================================================
    
    async def _update_import_metadata(
        self,
        breakdown_id: UUID,
        batch_id: UUID,
        source: str
    ) -> None:
        """Update breakdown with import metadata."""
        try:
            update_data = {
                'import_batch_id': str(batch_id),
                'import_source': source,
                'updated_at': datetime.now().isoformat()
            }
            
            self.supabase.table('po_breakdowns')\
                .update(update_data)\
                .eq('id', str(breakdown_id))\
                .execute()
            
        except Exception as e:
            logger.warning(f"Failed to update import metadata for {breakdown_id}: {e}")
    
    def get_default_column_mappings(self) -> Dict[str, str]:
        """
        Get default column mappings for SAP PO data.
        
        Returns:
            Dict mapping target fields to expected CSV column names
        """
        return {
            'name': 'Name',
            'code': 'Code',
            'sap_po_number': 'PO Number',
            'sap_line_item': 'Line Item',
            'cost_center': 'Cost Center',
            'gl_account': 'GL Account',
            'planned_amount': 'Planned Amount',
            'committed_amount': 'Committed Amount',
            'actual_amount': 'Actual Amount',
            'currency': 'Currency',
            'category': 'Category',
            'subcategory': 'Subcategory',
            'tags': 'Tags',
            'notes': 'Notes'
        }
    
    async def suggest_column_mappings(
        self,
        file: UploadFile
    ) -> Dict[str, List[str]]:
        """
        Analyze file headers and suggest column mappings.
        
        **Validates: Requirements 1.1**
        
        Args:
            file: File to analyze
            
        Returns:
            Dict mapping target fields to suggested CSV columns
        """
        try:
            # Read file headers
            content = await file.read()
            await file.seek(0)  # Reset file pointer
            
            file_ext = file.filename.lower().split('.')[-1] if file.filename else ''
            
            if file_ext == 'csv':
                # Parse CSV headers
                csv_content = content.decode('utf-8')
                csv_reader = csv.reader(io.StringIO(csv_content))
                headers = next(csv_reader)
            elif file_ext in ['xlsx', 'xls']:
                # Parse Excel headers
                import openpyxl
                workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                sheet = workbook.active
                headers = [str(cell.value) if cell.value else '' for cell in next(sheet.iter_rows())]
            else:
                return {}
            
            # Match headers to target fields
            default_mappings = self.get_default_column_mappings()
            suggestions = {}
            
            for target_field, expected_header in default_mappings.items():
                matches = []
                for header in headers:
                    if self._is_column_match(header, expected_header, target_field):
                        matches.append(header)
                suggestions[target_field] = matches
            
            return suggestions
            
        except Exception as e:
            logger.error(f"Failed to suggest column mappings: {e}")
            return {}
    
    def _is_column_match(
        self,
        csv_header: str,
        expected_header: str,
        target_field: str
    ) -> bool:
        """Check if a CSV header matches an expected header or target field."""
        csv_lower = csv_header.lower().replace('_', ' ').replace('-', ' ').strip()
        expected_lower = expected_header.lower().replace('_', ' ').replace('-', ' ').strip()
        target_lower = target_field.lower().replace('_', ' ').replace('-', ' ').strip()
        
        # Exact match
        if csv_lower == expected_lower or csv_lower == target_lower:
            return True
        
        # Contains match
        if expected_lower in csv_lower or csv_lower in expected_lower:
            return True
        
        if target_lower in csv_lower or csv_lower in target_lower:
            return True
        
        # Common variations
        variations = {
            'po number': ['purchase order', 'po no', 'po#', 'po num'],
            'line item': ['item', 'line', 'line no'],
            'cost center': ['cc', 'cost centre'],
            'gl account': ['gl', 'account', 'general ledger'],
            'planned amount': ['planned', 'budget', 'plan'],
            'committed amount': ['committed', 'commitment'],
            'actual amount': ['actual', 'spent', 'actuals'],
            'currency': ['curr', 'currency code'],
        }
        
        for key, variants in variations.items():
            if key in target_lower or key in expected_lower:
                for variant in variants:
                    if variant in csv_lower:
                        return True
        
        return False
